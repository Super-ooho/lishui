from openpyxl import load_workbook
#读取路径
book = load_workbook(filename=r"res.xlsx")
#读取名字为Sheet1的表
sheet = book.get_sheet_by_name("3倍水电")
#用于存储数据的数组
res = []

for i in range(1,28):
    row_num = 1
    data= []
    while row_num <= 21 :
        #将表中第一列的1-100行数据写入data数组中
        data.append(sheet.cell(row=row_num, column=i).value)
        row_num = row_num + 1
    res.append(data)
print(res)